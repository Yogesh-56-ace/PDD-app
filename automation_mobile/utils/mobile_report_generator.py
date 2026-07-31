import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill

class MobileReportGenerator:
    @staticmethod
    def generate_mobile_reports(test_results, duration):
        results_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'Test Results'))
        os.makedirs(os.path.join(results_dir, 'Excel'), exist_ok=True)
        os.makedirs(os.path.join(results_dir, 'HTML'), exist_ok=True)

        passed = [t for t in test_results if t['status'] == 'Passed']
        failed = [t for t in test_results if t['status'] == 'Failed']
        total = len(test_results)
        pass_rate = round((len(passed) / total * 100), 2) if total > 0 else 100.0

        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        # Excel Appium Report
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Appium Mobile E2E Results"
        headers = ["Test ID", "Module", "Test Name", "Priority", "Status", "Duration"]
        ws.append(headers)

        for t in test_results:
            ws.append([t['id'], t['module'], t['name'], t['priority'], t['status'], f"{t['duration']}s"])

        for col in range(1, 7):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).font = header_font

        excel_path = os.path.join(results_dir, 'Excel', 'Appium_Mobile_Automation_Report.xlsx')
        wb.save(excel_path)

        print(f"[SUCCESS] Generated Appium Mobile Excel Report with {total} test cases (Pass Rate: {pass_rate}%).")
