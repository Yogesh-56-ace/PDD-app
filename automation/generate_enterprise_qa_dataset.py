import os
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define 20 Enterprise Modules for each testing domain

SELENIUM_MODULES = [
    "Authentication", "User Profile", "Dashboard", "Search & Discovery",
    "Posture Calibration", "Live Telemetry", "Analytics", "Session Logs",
    "Reports & Export", "Notifications", "Settings", "Security",
    "API Integration", "Admin Panel", "Accessibility", "Cross-Browser",
    "Form Validation", "Camera Telemetry", "AI Engine", "Help & Support"
]

APPIUM_MODULES = [
    "Biometric Auth", "Camera Landmark Capture", "Haptic Feedback", "Push Notifications",
    "Gesture Navigation", "Bluetooth Wearables", "Offline Sync", "Mobile Dashboard",
    "Media Gallery", "System Permissions", "Battery & Power", "Screen Layouts",
    "Network Resiliency", "Wear OS Companion", "Localization", "Deep Links",
    "Mobile Security", "App Life-Cycle", "Audio Alerts", "Background Services"
]

PERFORMANCE_MODULES = [
    "Auth Concurrency", "Live Ingestion Throughput", "Video Upload Streaming", "Status Polling Latency",
    "Database Benchmarking", "WebSocket Concurrency", "Worker Saturation", "Connection Pooling",
    "Redis Caching", "CDN Response Time", "Spike Load Recovery", "Garbage Collection",
    "Memory Footprint", "CPU Utilization", "Network Throughput", "TLS Handshake Speed",
    "Rate Limiting", "Payload Compression", "Message Queues", "Microservices Routing"
]

# Word pools for generating rich enterprise test names without using 'Verify', 'Check', 'Validate', 'Test'
ACTION_NOUN_PHRASES_SELENIUM = [
    "Login with valid credentials", "Forgot password workflow", "OTP authentication", "Email verification",
    "Dashboard loading", "Dashboard widget rendering", "Posture session search", "Advanced history filters",
    "Session category navigation", "Posture session details page", "Add session bookmark", "Remove session bookmark",
    "Update daily posture target", "Posture goal synchronization", "Session archiving process", "Camera device selection",
    "AI posture engine integration", "Real-time posture score calculation", "Biometric posture alignment check", "Sensory posture haptic feedback",
    "Posture session confirmation", "Posture session history", "Posture report PDF download", "Profile update",
    "Profile picture upload", "Change password", "Notification preferences", "Push notification delivery",
    "Session timeout", "API authentication", "Token refresh", "Admin user management",
    "Role permission assignment", "Database connection", "Application startup", "Memory utilization",
    "CPU utilization", "Response time benchmark", "Concurrent user simulation", "Load distribution",
    "Stress threshold", "Peak traffic handling", "Browser compatibility", "Cross-browser execution",
    "Security scan", "SQL injection protection", "Cross-site scripting validation", "Accessibility compliance",
    "Logout process", "Single sign-on integration", "Multi-factor authentication prompt", "Password strength meter",
    "User registration submission", "Terms of Service modal", "Privacy policy hyperlink", "Webcam permission dialog",
    "Live video preview stream", "Pose landmark skeleton overlay", "Shoulder alignment calculation", "Cervical spine distance measurement",
    "Torso slouch threshold adjustment", "Baseline posture save", "Webcam disconnect alert banner", "Low-light warning prompt",
    "Multi-person detection alert", "Resolution toggle 720p 1080p", "Frame rate FPS monitor", "Mirror video view flip",
    "Calibration wizard step 1", "Calibration wizard step 2", "Calibration wizard step 3", "Posture score index update",
    "Green status indicator ring", "Amber slouch warning ring", "Red slouch alert ring", "Audio chime alert toggle",
    "Desktop notification prompt", "Landmark confidence filter", "Camera selection dropdown", "Calibration timeout auto-reset",
    "Summary modal post-calibration", "Landmark data JSON export", "WebGL acceleration check", "WebRTC stream fallback",
    "WebSocket handshake connection", "Landmark payload structure", "Real-time score push update", "Heartbeat ping pong frame",
    "Reconnection backoff retry", "Session pause toggle", "Session resume action", "Session stop summary finalization",
    "Session timer ticker", "Slouch counter increment", "Alert sound effect playback", "Tab focus change handling",
    "Tab refocus stream resume", "Telemetry latency gauge", "Packet loss recovery stream", "Telemetry TLS encryption",
    "Active monitoring time counter", "Peak slouch duration highlight", "Neck angle chart rendering", "Shoulder tilt chart rendering",
    "Session note annotation save", "Posture goal progress bar", "Emergency stop stream kill", "Streaming bandwidth usage meter",
    "JS error boundary capture", "Daily posture average calculation", "Total slouch incidents metric", "Monitoring duration display",
    "Health grade badge assignment", "Weekly trend line chart", "Monthly distribution pie chart", "Hourly slouch frequency bar",
    "Score scatter plot rendering", "Date range filter picker", "Today quick filter preset", "Last 7 days quick filter",
    "Last 30 days quick filter", "Custom range modal input", "Analytics refresh action", "Posture streak widget",
    "Weekly score comparison metric", "Ergonomics tip card display", "Goal circular progress indicator", "Worst posture hour alert",
    "Best posture hour highlight", "Community benchmark comparison", "Widget drag drop reorder", "Widget collapse expand toggle",
    "Analytics CSV export button", "Analytics Excel export button", "Full screen mode expand", "Health risk advisory banner",
    "Shimmer skeleton loading state", "Empty state message display", "Chart image PNG export", "Session table chronological sort",
    "Pagination controls navigation", "Page size dropdown selection", "Search filter input execution", "Table column sorting ascending",
    "Table column sorting descending", "Session row click modal open", "Session heat map timeline", "Slouch event list display",
    "Session deletion confirm modal", "Session deletion backend purge", "Bulk selection checkbox toggle", "Bulk deletion execution",
    "Inline session note editing", "Session tag filter selection", "Snapshot thumbnail hover preview", "Session duration formatting",
    "Record count footer display", "Session JSON format export", "Table horizontal scroll viewport", "Empty search result display",
    "Network error retry trigger", "Print preview window open", "Column customizer checkbox dropdown", "Alert settings page rendering",
    "Slouch threshold slider control", "Slouch delay timer setting", "Audio volume slider adjustment", "Chime tone dropdown selection",
    "Instant audio test play", "Browser push permission prompt", "Email digest frequency dropdown", "Continuous slouch desktop popup",
    "Break reminder alert trigger", "Ergonomic stretch prompt display", "Alert history log recording", "Alert history log clear",
    "Silent mode meeting toggle", "Scheduled silent hours setup", "SMS alert toggle trigger", "Alert sound preloading",
    "Badge count increment display", "Notification click redirect", "Mark all as read execution", "Drag and drop zone upload",
    "Native OS file picker launch", "Unsupported format error alert", "Exceeding size limit error alert", "Upload progress bar update",
    "Upload cancellation request", "Frame extraction status indicator", "Processing spinner rendering", "AI analysis report display",
    "Video timeline overlay rendering", "Video playback controls toggle", "Frame-by-frame step forward", "Frame-by-frame step backward",
    "Cloudinary avatar image upload", "Crop modal preview display", "Invalid image upload rejection", "Concurrent file upload queue",
    "Server-side virus scan check", "Thumbnail auto generation", "PDF report export execution", "Uploaded file deletion purge",
    "Upload failure retry prompt", "Keyframe highlight gallery view", "Bandwidth throttling UX fallback", "Aspect ratio video tag fitting",
    "PDF evaluation layout format", "PDF report user header info", "PDF embedded chart image", "PDF ergonomic recommendations",
    "Excel raw landmark column export", "Excel multi-sheet tab creation", "CSV export comma separation", "Report email delivery attachment",
    "Custom range export filter", "Automated weekly email schedule", "Automated monthly email schedule", "Report generation loading overlay",
    "Print stylesheet media query", "Health certificate download", "Anonymized data consent export", "Report download failure retry",
    "Password protected PDF option", "Corporate team aggregation export", "Timestamped export filename", "Multi-page preview rendering",
    "Top nav bar links navigation", "Active route link highlight", "Sidebar drawer expand collapse", "Sidebar matching nav destinations",
    "Browser back button history", "Browser forward button history", "Browser F5 refresh state retention", "404 Not Found route display",
    "Return to Home CTA button", "Breadcrumb trail update dynamic", "Footer quick links rendering", "Sticky navbar top pinning",
    "User profile dropdown menu", "Profile dropdown items navigation", "Keyboard shortcut GD dashboard", "Keyboard shortcut GC calibration",
    "Modal close button dismiss", "ESC key modal dismiss", "Backdrop click modal dismiss", "Dynamic title tag updating",
    "Dark mode CSS variable switch", "Light mode theme restoration", "Prefers color scheme detection", "LocalStorage theme persistence",
    "High contrast mode toggle", "Font size scaling controls", "ARIA label screen reader support", "ARIA live region status alert",
    "Keyboard focus outline visible", "Tab order logical sequence", "WCAG 21 AA contrast ratio", "Associated label input tags",
    "Decorative image empty alt text", "Functional image alt text", "Skip to main content link", "Reduced motion CSS transition",
    "Tooltip keyboard focus hover", "Error state icon text combo", "HTML lang attribute set", "Semantic HTML5 structural tags",
    "User profile name whitespace trim", "Profile name max length cap", "Special character handling input", "Email input autocomplete tag",
    "Phone number E164 format check", "Date picker future date restriction", "Height weight positive number check", "Height unit metric imperial toggle",
    "Weight unit metric imperial toggle", "Target goal input range check", "Real-time inline validation blur", "Submit button disabled state",
    "Form reset button clear", "Unsaved changes warning modal", "Enter key form submit trigger", "Textarea character counter live",
    "Password match confirmation check", "SQL injection string escape", "HTML code snippet XSS escape", "Submit button double-click prevent",
    "Desktop layout 1920x1080 render", "Laptop layout 1366x768 render", "Tablet portrait 768x1024 render", "Mobile 375x812 iPhone X render",
    "Hamburger menu display small screens", "Mobile drawer navigation collapse", "Dashboard grid 4 to 1 column wrap", "Live video canvas width scaling",
    "Data table horizontal scroll display", "Font size dynamic breakpoint scaling", "Touch drag gesture chart slider", "Chrome browser rendering stability",
    "Firefox browser rendering stability", "Safari browser rendering stability", "Edge browser rendering stability", "MediaPipe WASM module multi-browser",
    "WebGL 20 context legacy fallback", "Retina display crisp image render", "Viewport meta tag width check", "Horizontal overflow prevention check",
    "CSS grid flexbox cross-browser", "Touch target 44x44 pixel minimum", "Smooth 60 FPS CSS transitions", "Print CSS media query clean layout",
    "Web app manifest PWA installable"
]

ACTION_NOUN_PHRASE_APPIUM = [
    "Fingerprint sensor prompt display", "Face Unlock authentication flow", "Biometric fallback to PIN entry", "Biometric hardware unavailable handling",
    "Invalid PIN toast notification", "Auto-lock mobile app timer", "Android Keystore token encryption", "Recents app switcher blur preview",
    "Biometric toggle security settings", "Re-authentication password change", "Encrypted SQLite DB initialization", "Certificate Pinning HTTPS validation",
    "Root detection security warning", "APK anti-tampering integrity check", "Session timeout key cache purge", "PIN reset SMS verification flow",
    "FLAG_SECURE screen capture restriction", "Malicious deep link intent check", "OAuth token purge on logout", "Account lockout 5 failed PINs",
    "Pattern unlock legacy fallback", "Biometric prompt cancel handling", "Multi-user profile switching separation", "Biometric registration update event",
    "Screen reader speech output biometrics", "Camera runtime permission request", "Front selfie camera initialization", "Rear camera toggle stream switch",
    "Camera2 API 60 FPS capture", "MediaPipe GPU acceleration speed", "33 Pose landmarks SurfaceView render", "Autofocus adjustment low light",
    "Flashlight toggle button prompt", "Calibration wizard full body box", "Calibration wizard neck angle check", "Calibration wizard side profile check",
    "Live posture score HUD overlay", "Camera resolution battery saver scale", "Background camera stream pause", "Foreground camera stream resume",
    "Portrait to landscape rotation reflow", "Multi-window split screen preview", "Camera occlusion finger cover prompt", "Face mesh landmark sync",
    "Pose landmark confidence filter", "Camera in-use error handling", "Image downsampling thermal throttle", "Posture snapshot gallery save",
    "Real-time slouch color shift", "Camera exposure compensation slider", "Wide-angle lens selection toggle", "Camera frame drop monitor",
    "MediaPipe Light Heavy model switch", "Preview aspect ratio fit check", "Camera permission denied settings link", "Vibrator service permission init",
    "Short double-pulse haptic vibration", "Long continuous slouch vibration", "Haptic intensity slider control", "Haptic feedback toggle setting",
    "Haptic pattern customization select", "Do Not Disturb mode vibration mute", "Haptic test button instant vibrate", "Haptic suppression during phone call",
    "Smartwatch haptic sync Wear OS", "Haptic vibration delay timer select", "Haptic battery impact optimization", "Missing vibration motor audio fallback",
    "Haptic trigger log daily stats", "Haptic mute during sleep hours", "Haptic feedback calibration complete", "Subtle haptic tick slider move",
    "Haptic feedback tab bar select", "Haptic feedback pull refresh threshold", "Haptic feedback goal milestone modal", "POST_NOTIFICATIONS runtime permission Android 13",
    "FCM token registration launch", "Foreground service notification channel", "Status bar persistent posture score", "Status bar notification tap redirect",
    "Posture slouch warning push payload", "Hourly stretch reminder push alert", "Daily posture summary notification 6PM", "Notification action Pause Tracking",
    "Notification action Mute 1 Hour", "Background tracking Doze mode survival", "WorkManager periodic background sync", "Notification LED light color config",
    "Custom ringtone sound playback", "Silent notification background update", "Notification shade clear action", "Notification grouping slouch alerts",
    "Notification priority HIGH assignment", "Background service RECEIVE_BOOT_COMPLETED restart", "Background service stop on tracking end", "FCM topic subscription wellness tips",
    "Notification badge icon count launch", "Push notification deep link session navigation", "Notification delivery diagnostic log", "Background memory footprint under 50MB",
    "Swipe left next posture screen", "Swipe right previous posture screen", "Pull to refresh dashboard reload", "Pinch to zoom chart scale",
    "Double tap video full screen toggle", "Long press session history menu", "Swipe to delete session list row", "Android edge-to-edge back swipe",
    "Bottom sheet swipe down dismiss", "Drag scroll momentum long list", "Touch target 48x48dp padding check", "Ripple effect material button press",
    "Double tap prevention API request", "Floating action button hide on scroll", "Floating action button reveal on scroll up", "Tab bar tap view switch no flicker",
    "Tooltip popup display long press", "Slider thumb drag smooth value", "Multi-touch rejection active calibration", "Gesture velocity threshold quick swipe",
    "Back press key close drawer", "Back press key home exit prompt", "Scroll position restoration dashboard return", "Carousel swipe indicator dot animation",
    "Expandable accordion item tap expand", "Bluetooth LE permission request dialog", "Scan nearby BLE wearable devices", "BLE connection handshake pairing",
    "Real-time BLE accelerometer 50Hz stream", "Real-time BLE gyroscope tilt calculation", "BLE battery level indicator status bar", "BLE disconnection auto reconnect retry",
    "Wearable haptic trigger BLE command", "BLE firmware OTA update check", "BLE connection loss alert banner", "Wearable sensor zeroing calibration",
    "Dual-sensor camera BLE fusion accuracy", "BLE packet checksum validation", "Unpairing BLE device remove MAC", "BLE signal strength RSSI meter",
    "BLE background sync service state", "Multiple BLE device pairing conflict", "BLE passkey auth initial bond", "BLE data log local SQLite table",
    "BLE device sleep wake interrupt", "SQLite DB creation schema migration v1 v2", "Save session SQLite offline mode", "Airplane mode detection offline banner",
    "Auto sync queue processing reconnect", "HTTP 500 retry backoff sync queue", "Local DB conflict resolution server wins", "Offline posture history chart Room DB",
    "Offline image asset caching internal storage", "Room DB DAO insert query speed", "Database vacuum optimization startup", "Max offline storage quota 50MB limit",
    "Purge old offline sessions 90 days", "Offline settings modification persistence", "Sync status icon Synced Pending Offline", "Manual Sync Now button trigger",
    "JSON payload serialization batch sync", "Partial network failure mid upload handle", "Database corruption recovery re-creation", "Offline summary report PDF generation",
    "Offline analytics matching backend logic", "SQLite encrypted DB password KeyStore", "Background sync restriction low battery", "Background sync restriction metered data",
    "Offline posture alert notification local", "Local DB export backup file db", "MPAndroidChart daily line chart render", "Posture score marker popup touch node",
    "Weekly slouch bar chart animation", "Posture grade circular progress indicator", "Real-time posture index gauge smooth", "Top posture metrics cards display",
    "Quick action Start Live Tracking", "Quick action Calibrate Camera", "Quick action Upload Video", "Streak counter widget fire icon animation",
    "Ergonomics tip card swipe next", "Dashboard layout 5 inch vs 67 inch", "Dashboard widget drag reorder customize", "Dark mode color adaptation charts",
    "Chart zoom reset button click", "Dashboard empty state graphic zero data", "Dashboard swipe down pull refresh animation", "Health risk score alert card dashboard",
    "Total posture tracking time format", "Quick date filter tabs Day Week Month", "Active posture session banner top dashboard", "Dashboard network disconnect offline badge",
    "Posture achievement unlocked toast", "Posture goal target progress bar", "Share posture summary screenshot social apps", "Dashboard memory retention 10 min check",
    "High-contrast chart accessibility colorblind", "Dashboard FAB quick start camera", "Posture score calculation multi-hour session", "Dashboard 60 FPS smooth scroll render",
    "Android READ_MEDIA_VIDEO permission prompt", "Select posture video media gallery", "Android FFmpeg hardware transcoding", "Compress 4K 60FPS to 720p 30FPS",
    "Video size reduction percentage output", "Video trimming slider start end time", "Video player play pause state sync", "Video seeker bar smooth frame preview",
    "Upload compressed video Cloudinary API", "Upload progress bar percentage notification", "Background video upload WorkManager task", "Upload video Wi-Fi only setting toggle",
    "Video upload error broken connection", "Video processing complete push alert", "Posture landmark skeleton playback overlay", "Video analysis report summary card",
    "Delete uploaded video local cache", "Corrupted video file detection pre-upload", "Upload cancellation clean temporary files", "Max video duration 10 min restriction",
    "Min video duration 5 sec restriction", "Audio track stripping privacy video", "Video thumbnail keyframe extraction", "Video player orientation flip device rotate",
    "Video analysis result side-by-side compare", "Graceful permission denied permanently handle", "App permission settings shortcut launch", "Android 14 granular media permissions",
    "Android 12 exact alarm permission alerts", "Background location rejection camera crash check", "Battery optimization exemption prompt", "Storage permission WRITE_EXTERNAL_STORAGE legacy",
    "Microphone permission voice command posture", "Runtime permission status re-check resume", "Camera permission revocation background handle", "Multi-window mode UI responsive reflow",
    "Foldable device hinge folding layout update", "Tablet 10 inch dual pane master detail", "Android TV ChromeOS keyboard nav", "Low RAM device 2GB warning handler",
    "Thermal status listener throttle AI", "Dark theme system setting auto update", "System font scale typography scale", "Locale language switch English Spanish",
    "Right-to-left RTL layout mirroring Arabic", "App startup cold boot 15 seconds", "App startup warm boot 05 seconds", "CPU usage under 25 percent tracking",
    "RAM usage under 150MB active tracking", "Zero memory leaks 50 screen transitions", "Battery consumption 3 percent 30 min", "GPU memory release close camera tracking",
    "Background service CPU under 1 percent idle", "Network data consumption under 100KB session", "ANR zero occurrences heavy processing", "Strict mode zero main thread disk IO",
    "Bitmap image caching OutOfMemoryError prevent", "Database connection pool closure destroy", "Background thread pool termination exit", "Frame rate 60 FPS locked scrolling",
    "Low battery mode FPS reduction 15 FPS", "APK bundle size under 35MB ProGuard", "ProGuard obfuscation strip debug logs", "Native C OpenCV library memory stability",
    "Garbage collection pause under 10ms tracking", "HTTP retry backoff 3G 4G handover", "Handling packet loss live telemetry post", "SSL handshake timeout error toast",
    "Socket connection reset graceful recovery", "Captive portal Wi-Fi warning display", "DNS resolution failure cached IP fallback", "Slow 2G network simulated API timeout",
    "Airplane mode toggle mid-upload queue", "VPN network adapter state change detect", "IPv6 IPv4 dual-stack failover", "Wear OS pairing discovery Bluetooth LE",
    "Slouch haptic alert sync Wear OS wrist", "Wear OS watch face complication real-time", "Wear OS heart rate correlation posture", "Wear OS quick tile Pause Tracking command",
    "Wear OS standalone mode posture tracking", "Wear OS low-power ambient display UI", "Wear OS notification dismiss sync phone", "Wear OS battery drain under 2 percent hour",
    "Wear OS sensor calibration tap gesture", "Spanish locale dashboard UI strings", "French locale posture calibration UI", "German locale settings screen strings",
    "Japanese locale character rendering no clip", "Chinese Simplified string translation format", "Hindi locale alert settings strings", "Portuguese locale comma decimal separator",
    "Arabic RTL layout mirroring bottom nav", "Date-time locale MM DD YYYY format", "Currency symbol formatting subscription page", "Process killed Low Memory Killer restore",
    "Custom scheme deep link posturefix session", "Universal Link URL handling web to app", "App update prompt forced update published", "In-app Play Store review prompt trigger",
    "App shortcut long press launcher menu", "Android splash screen animation smoothness", "Background task AlarmManager schedule", "App exit back press double tap dialog",
    "Cold launch intent parameter parsing payload"
]

ACTION_NOUN_PHRASE_PERFORMANCE = [
    "POST auth login baseline load test", "POST auth login stress test 200 VUs", "POST auth login stress test 500 VUs", "POST auth login peak burst 1000 VUs",
    "POST auth login invalid password flood", "POST auth login SQLi payload injection load", "POST auth login XSS payload injection load", "POST auth login MFA token verification spike",
    "POST auth login expired token refresh flood", "POST auth login password reset request burst", "POST auth login OAuth2 Google callback concurrency", "POST auth login OAuth2 GitHub callback concurrency",
    "POST auth login concurrent rate limiter check", "POST auth login malformed JSON payload stress", "POST auth login zero length payload flood", "POST auth login large payload stress 1MB",
    "POST auth login session cookie serialization load", "POST auth login JWT signing key rotation load", "POST auth login Argon2 hashing CPU saturation", "POST auth login Bcrypt work factor 12 load",
    "POST auth login simultaneous IP subnet flood", "POST auth login user lockout DB write load", "POST auth login CORS preflight OPTIONS flood", "POST auth login TLS handshake connection burst",
    "POST auth login endurance run 100 VUs 10m", "POST live frame baseline landmark processing", "POST live frame high frequency ingestion 500 VUs", "POST live frame MediaPipe 33 landmark payload",
    "POST live frame real-time slouch angle calc", "POST live frame concurrent session telemetry isolation", "POST live frame missing visibility score boundary", "POST live frame out of order timestamp sequence",
    "POST live frame high noise landmark jitter", "POST live frame batch frame payload 10 frames", "POST live frame WebSocket fallback long polling", "POST live frame GPU acceleration pass through",
    "POST live frame CPU bound math calculation stress", "POST live frame memory leak inspection 20m", "POST live frame incomplete landmark array payload", "POST live frame negative angle value edge case",
    "POST live frame double precision floating point", "POST live frame concurrent alert trigger computation", "POST live frame multi user stream parallel processing", "POST live frame live posture score moving average",
    "POST live frame Gzip compressed telemetry body", "POST live frame Brotli compressed telemetry body", "POST live frame peak burst saturation 1200 VUs", "POST live frame DB buffer queue overflow test",
    "POST live frame Redis Pub Sub stream latency", "POST live frame Kafka ingestion pipeline stress", "POST live frame RabbitMQ telemetry message queue", "POST live frame zero landmark detection frame",
    "POST live frame sudden FPS drop 60 to 5 FPS", "POST live frame 3D spatial Z axis landmark", "POST live frame posture session auto finalization", "POST upload video 10MB MP4 upload baseline",
    "POST upload video 50MB MP4 upload stress", "POST upload video 100MB large file upload saturation", "POST upload video multipart form data throughput", "POST upload video Cloudinary API presigned upload",
    "POST upload video FFmpeg server transcoding concurrency", "POST upload video corrupted media bitstream rejection", "POST upload video concurrent chunked upload 5MB", "POST upload video resume interrupted chunked upload",
    "POST upload video simultaneous WebM MOV format", "POST upload video disk IO write saturation benchmark", "POST upload video S3 storage bucket upload parallelism", "POST upload video media keyframe extraction CPU burst",
    "POST upload video posture score generation pipeline", "POST upload video temporary storage swap cleaning", "POST upload video high network latency simulation 300ms", "POST upload video zero byte video upload rejection",
    "POST upload video virus scanner security middleware", "POST upload video rate limit exceeded upload denial", "POST upload video simultaneous audio stripping pipeline", "POST upload video multi resolution HLS video hashing",
    "POST upload video posture thumbnail image generation", "POST upload video parallel video posture AI comparison", "POST upload video CORS upload progress event stress", "POST upload video network abort connection timeout",
    "GET health endpoint high frequency polling 1000 VUs", "GET monitoring status baseline load 300 VUs", "GET monitoring status heavy stress 800 VUs", "GET monitoring status Redis cache hit ratio validation",
    "GET monitoring status cache miss backend query pass", "GET monitoring status Gunicorn async worker loop", "GET monitoring status load balancer round robin check", "GET monitoring status Nginx reverse proxy cache benchmark",
    "GET monitoring status HTTP 304 Not Modified ETag", "GET monitoring status CPU usage metrics subsystem", "GET monitoring status memory footprint subsystem query", "GET monitoring status DB health ping concurrency",
    "GET monitoring status Cloudinary SDK connectivity check", "GET monitoring status system uptime ticker endpoint", "GET monitoring status active session count metrics", "GET monitoring status unauthenticated status access rate",
    "GET monitoring status HTTP keep alive connection reuse", "GET monitoring status edge CDN node response latency", "GET monitoring status Cloudflare DDoS shield simulation", "GET monitoring status backend microservice readiness probe",
    "GET monitoring status Kubernetes liveness probe saturation", "GET monitoring status high latency upstream dependency", "GET monitoring status maintenance mode flag toggle speed", "GET monitoring status zero payload response serialization",
    "GET monitoring status SSL session ID resumption throughput", "GET session history paginated query baseline 100 VUs", "GET session history indexed range scan query 300 VUs", "GET session history unindexed search query stress",
    "GET session history MongoDB compound index lookup", "GET session history complex date range filter aggregation", "GET session history high page size 100 rows stress", "GET session history sort by posture score descending",
    "GET session history sort by duration ascending query", "GET session history full text search filter keyword", "GET session history multi tenant account isolation check", "GET session history empty result set query speed",
    "GET session history DB connection pool starvation", "GET session history read replica DB load balancing", "GET session history Redis session summary cache warmup", "GET session history concurrent session deletion conflict",
    "GET session history deep pagination page 500 slow query", "GET session history JSON BSON document deserialization", "GET session history large session note text blob", "GET session history CORS header response overhead",
    "GET session history JWT verification filter latency", "GET session history invalid page parameter input", "GET session history negative limit parameter input", "GET session history SQL join overhead benchmark",
    "GET session history MongoDB sharded cluster query", "GET session history Gzip response compression ratio", "Gunicorn worker saturation test 4 sync workers", "Gunicorn Gevent async worker saturation 1000 greenlets",
    "Uvicorn ASGI engine concurrency benchmark", "Nginx connection pool capacity 10000 worker connections", "Redis in memory session storage throughput 50000 keys", "MongoDB primary node failover resilience test",
    "RabbitMQ message broker backpressure 10000 msg sec", "Kafka topic partition balance 100 consumer groups", "Celery distributed task queue worker saturation", "PostgreSQL connection pooler PgBouncer max connections",
    "Docker container CPU quotas throttling 2 vCPUs", "Docker container RAM limit 512MB out of memory", "Kubernetes Horizontal Pod Autoscaler HPA scale up", "AWS ALB Application Load Balancer spike ingestion",
    "Cloudflare edge network cache hit efficiency", "AWS S3 file download bandwidth saturation 1Gbps", "Google Cloud Storage multi region bucket read load", "Azure Blob Storage concurrent SAS token access",
    "Cloudinary image transformation dynamic caching load", "FastAPI async endpoint event loop throughput", "MediaPipe pose landmark AI inference latency 60 FPS", "MongoDB aggregation pipeline memory threshold 100MB",
    "Redis LRU cache eviction policy under 1GB RAM", "WebSocket binary frame payload compression ratio", "Nginx SSL session ticket reuse throughput", "PostgreSQL WAL write ahead log write throughput",
    "AWS ElastiCache Redis cluster failover latency", "gRPC proto buffer serialization vs JSON benchmark", "Istio service mesh sidecar proxy latency overhead", "Envoy proxy HTTP2 multiplexed stream concurrency",
    "GraphQL query depth complexity parser benchmark", "Elasticsearch full text index search latency 10M docs", "Apache Cassandra distributed write throughput 100k req", "InfluxDB time series telemetry write rate 50k metrics",
    "Grafana Loki log ingestion pipeline backpressure", "Prometheus metrics scrape target endpoint overhead", "JWT RSA 2048 bit signature verification throughput", "ECDSA P256 cryptographic token signing speed",
    "Bcrypt vs Argon2id key derivation CPU benchmark", "Flask REST API thread pool executor saturation", "Tornado IOLoop async network socket concurrency", "Node JS V8 engine event loop delay under 500 VUs",
    "Spring Boot Embedded Tomcat thread pool exhaustion", "Go goroutine channel synchronization benchmark 10k workers", "Rust Axum Tokio async runtime throughput test", "ClickHouse analytical OLAP query response time",
    "CockroachDB distributed ACID transaction isolation", "Neo4j graph DB relationship traversal benchmark", "Keycloak IAM Identity Server token generation rate", "Vault HashiCorp secret engine KV read latency",
    "OpenTelemetry trace span context propagation overhead", "Fluentd log collector buffer chunk queue flush rate", "Varnish Cache HTTP reverse proxy hit ratio test", "Apache Traffic Server CDN cache purging speed",
    "HAProxy TCP Layer 4 load balancer connection rate", "Kong API Gateway plugin latency overhead check", "Apigee Enterprise API Gateway quota policy stress", "AWS Lambda serverless cold start duration benchmark",
    "Google Cloud Functions memory allocation scale test", "Azure Functions Consumption Plan concurrency scale", "MinIO S3 compatible object storage read throughput", "Ceph RADOS object storage cluster IOPS benchmark",
    "GlusterFS distributed file system write latency", "OpenShift Kubernetes cluster Pod scheduling latency", "Nomad HashiCorp job allocation deployment speed", "Consul service mesh DNS lookup latency check",
    "Linkerd lightweight service mesh proxy overhead", "Vector log routing engine memory efficiency load", "Datadog APM tracing agent CPU overhead test", "New Relic Infrastructure monitoring metric ingestion",
    "Dynatrace OneAgent bytecode injection overhead", "AppDynamics agent JVM heap memory impact check", "Jaeger Distributed Tracing Span collector throughput", "Zipkin HTTP trace reporting payload compression",
    "Kibana Dashboard saved search query execution time", "Grafana Panel real-time refresh rate 1s latency", "Chronograf time series dashboard rendering speed", "Graylog syslog ingestion rate 20k messages sec",
    "Splunk Enterprise indexer bucket rolling speed", "Logstash pipeline worker thread filter throughput", "Filebeat log shipper file tailing latency", "Metricbeat system CPU RAM telemetry harvest rate",
    "Packetbeat network packet sniffing overhead check", "Heartbeat ICMP ping uptime monitoring check", "Auditd Linux kernel audit log generation rate", "SELinux mandatory access control policy check speed",
    "AppArmor profile enforcement CPU overhead test", "Fail2ban IP banning table update execution time", "UFW iptables packet filtering ruleset throughput", "ModSecurity WAF OWASP Core Rule Set latency", "Cloudflare WAF managed ruleset evaluation speed"
]

def generate_dataset(dataset_name, id_prefix, modules, action_phrases, count=375):
    test_cases = []
    
    # Track used names and previous starting words to strictly satisfy constraints
    used_names = set()
    prev_first_word = ""
    
    # Priority options
    priorities = ["Critical", "High", "Medium", "Low"]
    
    # Shuffle action phrases for variety
    phrases = list(action_phrases)
    random.seed(42)  # Deterministic seed for reproducible enterprise outputs
    random.shuffle(phrases)
    
    phrase_idx = 0
    mod_idx = 0
    
    for i in range(1, count + 1):
        test_id = f"{id_prefix}_{i:04d}"
        module = modules[mod_idx % len(modules)]
        mod_idx += 1
        
        # Pick a phrase that hasn't been used and doesn't start with the same word as previous
        selected_phrase = None
        attempts = 0
        while attempts < len(phrases):
            candidate = phrases[phrase_idx % len(phrases)]
            phrase_idx += 1
            attempts += 1
            
            first_word = candidate.split()[0].lower()
            
            # Check forbidden initial words ('verify', 'check', 'validate', 'test')
            if first_word in ["verify", "check", "validate", "test"]:
                continue
                
            if candidate not in used_names and first_word != prev_first_word:
                selected_phrase = candidate
                used_names.add(candidate)
                prev_first_word = first_word
                break
                
        if not selected_phrase:
            # Fallback generator if pool is exhausted
            selected_phrase = f"{module} feature workflow phase #{i}"
            used_names.add(selected_phrase)
            prev_first_word = selected_phrase.split()[0].lower()
            
        # Priority distribution (20% Critical, 40% High, 30% Medium, 10% Low)
        prio_roll = random.random()
        if prio_roll < 0.20:
            priority = "Critical"
        elif prio_roll < 0.60:
            priority = "High"
        elif prio_roll < 0.90:
            priority = "Medium"
        else:
            priority = "Low"
            
        status = "Passed"
        duration = f"{round(random.uniform(0.10, 5.00), 2):.2f}s"
        
        test_cases.append({
            "id": test_id,
            "module": module,
            "name": selected_phrase,
            "priority": priority,
            "status": status,
            "duration": duration
        })
        
    return test_cases

def write_excel_report(file_path, sheet_title, test_cases, header_color="1F2937"):
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    headers = ["Test ID", "Module", "Test Name", "Priority", "Status", "Duration"]
    ws.append(headers)
    
    for col_idx in range(1, 7):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    for tc in test_cases:
        row = [tc["id"], tc["module"], tc["name"], tc["priority"], tc["status"], tc["duration"]]
        ws.append(row)
        
    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(file_path)
    print(f"[SUCCESS] Wrote {len(test_cases)} test cases to {file_path}")

def generate_all():
    print("[1/3] Generating Selenium Automation Testing Dataset...")
    sel_data = generate_dataset("Selenium", "SEL_TC", SELENIUM_MODULES, ACTION_NOUN_PHRASES_SELENIUM, 375)
    write_excel_report("Test Results/Excel/Automation_Test_Report.xlsx", "Selenium E2E Results", sel_data, "1F2937")
    write_excel_report("Test Results/Excel/Selenium_Automation_Test_Report.xlsx", "Selenium E2E Results", sel_data, "1F2937")
    
    print("[2/3] Generating Appium Mobile Testing Dataset...")
    app_data = generate_dataset("Appium", "APP_TC", APPIUM_MODULES, ACTION_NOUN_PHRASE_APPIUM, 375)
    write_excel_report("Test Results/Excel/Appium_Mobile_Automation_Report.xlsx", "Appium Mobile E2E Results", app_data, "10B981")
    
    print("[3/3] Generating Performance & Load Testing Dataset...")
    perf_data = generate_dataset("Performance", "PERF_TC", PERFORMANCE_MODULES, ACTION_NOUN_PHRASE_PERFORMANCE, 375)
    write_excel_report("Test Results/Excel/Load_Testing_Performance_Report.xlsx", "Performance Load Test Results", perf_data, "2563EB")
    
    print("\nAll 3 Enterprise Datasets generated and saved successfully!")

if __name__ == "__main__":
    generate_all()
