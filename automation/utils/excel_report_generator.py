import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import Config
from automation.utils.logger import TestLogger

logger = TestLogger.get_logger()

class ExcelReportGenerator:
    @staticmethod
    def generate_all_excel_reports(test_results):
        """
        Generates 4 distinct Excel reports:
        1. Automation_Test_Report.xlsx (6 sheets)
        2. Failed_Test_Cases.xlsx
        3. Passed_Test_Cases.xlsx
        4. Summary_Report.xlsx
        """
        Config.ensure_directories()
        
        # Color definitions
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        pass_font = Font(name="Calibri", size=10, bold=True, color="166534")
        
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        fail_font = Font(name="Calibri", size=10, bold=True, color="991B1B")

        skip_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        skip_font = Font(name="Calibri", size=10, bold=True, color="92400E")

        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Failure Reason"]

        # ----------------------------------------------------
        # 1. Main Report: Automation_Test_Report.xlsx
        # ----------------------------------------------------
        wb_main = openpyxl.Workbook()
        
        # Sheet 1: Executed Test Cases
        ws_all = wb_main.active
        ws_all.title = "Executed Test Cases"
        
        # Sheet 2: Passed Tests
        ws_passed = wb_main.create_sheet(title="Passed Tests")
        
        # Sheet 3: Failed Tests
        ws_failed = wb_main.create_sheet(title="Failed Tests")
        
        # Sheet 4: Skipped Tests
        ws_skipped = wb_main.create_sheet(title="Skipped Tests")

        # Add headers to sheets 1-4
        for ws in [ws_all, ws_passed, ws_failed, ws_skipped]:
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

        passed_count = 0
        failed_count = 0
        skipped_count = 0
        total_time = 0.0

        module_metrics = {}

        for item in test_results:
            tid = item["test_id"]
            module = item["module"]
            name = item["test_name"]
            status = item["status"].upper()
            exec_time = float(item.get("duration", 0.1))
            priority = item.get("priority", "P2")
            reason = item.get("failure_reason", "")

            total_time += exec_time

            # Update module metrics
            if module not in module_metrics:
                module_metrics[module] = {"total": 0, "pass": 0, "fail": 0, "skip": 0}
            module_metrics[module]["total"] += 1

            row_data = [tid, module, name, status, round(exec_time, 3), priority, reason]
            ws_all.append(row_data)

            if status == "PASSED" or status == "PASS":
                passed_count += 1
                module_metrics[module]["pass"] += 1
                ws_passed.append(row_data)
            elif status == "FAILED" or status == "FAIL":
                failed_count += 1
                module_metrics[module]["fail"] += 1
                ws_failed.append(row_data)
            else:
                skipped_count += 1
                module_metrics[module]["skip"] += 1
                ws_skipped.append(row_data)

        # Style rows in sheets 1-4
        for ws in [ws_all, ws_passed, ws_failed, ws_skipped]:
            for row in range(2, ws.max_row + 1):
                st = ws.cell(row=row, column=4).value
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    if col in [1, 4, 5, 6]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply status styling
                status_cell = ws.cell(row=row, column=4)
                if st in ["PASSED", "PASS"]:
                    status_cell.fill = pass_fill
                    status_cell.font = pass_font
                elif st in ["FAILED", "FAIL"]:
                    status_cell.fill = fail_fill
                    status_cell.font = fail_font
                else:
                    status_cell.fill = skip_fill
                    status_cell.font = skip_font

            # Auto-fit columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Sheet 5: Execution Metrics
        ws_metrics = wb_main.create_sheet(title="Execution Metrics")
        ws_metrics.append(["Metric", "Value"])
        ws_metrics.append(["Total Executed", len(test_results)])
        ws_metrics.append(["Passed", passed_count])
        ws_metrics.append(["Failed", failed_count])
        ws_metrics.append(["Skipped", skipped_count])
        pass_rate = round((passed_count / max(len(test_results), 1)) * 100, 2)
        ws_metrics.append(["Pass Rate (%)", f"{pass_rate}%"])
        ws_metrics.append(["Total Duration (s)", round(total_time, 2)])

        for row in range(1, ws_metrics.max_row + 1):
            for col in range(1, 3):
                cell = ws_metrics.cell(row=row, column=col)
                cell.border = thin_border
                if row == 1:
                    cell.fill = header_fill
                    cell.font = header_font

        # Sheet 6: Defect Summary
        ws_defects = wb_main.create_sheet(title="Defect Summary")
        ws_defects.append(["Module", "Total Tests", "Passed", "Failed", "Skipped", "Module Pass Rate (%)"])
        for col in range(1, 7):
            c = ws_defects.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
        
        for mod, counts in module_metrics.items():
            tot = counts["total"]
            p = counts["pass"]
            f = counts["fail"]
            s = counts["skip"]
            m_rate = round((p / max(tot, 1)) * 100, 2)
            ws_defects.append([mod, tot, p, f, s, f"{m_rate}%"])

        for row in range(2, ws_defects.max_row + 1):
            for col in range(1, 7):
                ws_defects.cell(row=row, column=col).border = thin_border

        main_path = os.path.join(Config.EXCEL_DIR, "Automation_Test_Report.xlsx")
        wb_main.save(main_path)
        logger.info(f"Generated main Excel report: {main_path}")

        # ----------------------------------------------------
        # 2. Failed_Test_Cases.xlsx
        # ----------------------------------------------------
        wb_fail_only = openpyxl.Workbook()
        ws_fo = wb_fail_only.active
        ws_fo.title = "Failed Test Cases"
        ws_fo.append(headers)
        for col in range(1, len(headers) + 1):
            ws_fo.cell(row=1, column=col).fill = header_fill
            ws_fo.cell(row=1, column=col).font = header_font

        for item in test_results:
            if item["status"].upper() in ["FAILED", "FAIL"]:
                ws_fo.append([
                    item["test_id"], item["module"], item["test_name"],
                    "FAILED", round(float(item.get("duration", 0.1)), 3),
                    item.get("priority", "P2"), item.get("failure_reason", "")
                ])
        
        fail_path = os.path.join(Config.EXCEL_DIR, "Failed_Test_Cases.xlsx")
        wb_fail_only.save(fail_path)

        # ----------------------------------------------------
        # 3. Passed_Test_Cases.xlsx
        # ----------------------------------------------------
        wb_pass_only = openpyxl.Workbook()
        ws_po = wb_pass_only.active
        ws_po.title = "Passed Test Cases"
        ws_po.append(headers)
        for col in range(1, len(headers) + 1):
            ws_po.cell(row=1, column=col).fill = header_fill
            ws_po.cell(row=1, column=col).font = header_font

        for item in test_results:
            if item["status"].upper() in ["PASSED", "PASS"]:
                ws_po.append([
                    item["test_id"], item["module"], item["test_name"],
                    "PASSED", round(float(item.get("duration", 0.1)), 3),
                    item.get("priority", "P2"), ""
                ])

        pass_path = os.path.join(Config.EXCEL_DIR, "Passed_Test_Cases.xlsx")
        wb_pass_only.save(pass_path)

        # ----------------------------------------------------
        # 4. Summary_Report.xlsx
        # ----------------------------------------------------
        wb_sum = openpyxl.Workbook()
        ws_s = wb_sum.active
        ws_s.title = "Summary Report"
        ws_s.append(["Metric Name", "Value"])
        ws_s.cell(row=1, column=1).fill = header_fill
        ws_s.cell(row=1, column=1).font = header_font
        ws_s.cell(row=1, column=2).fill = header_fill
        ws_s.cell(row=1, column=2).font = header_font

        ws_s.append(["Target URL", Config.BASE_URL])
        ws_s.append(["Total Test Cases", len(test_results)])
        ws_s.append(["Passed", passed_count])
        ws_s.append(["Failed", failed_count])
        ws_s.append(["Skipped", skipped_count])
        ws_s.append(["Pass Percentage", f"{pass_rate}%"])
        ws_s.append(["Total Duration (Seconds)", round(total_time, 2)])

        summary_path = os.path.join(Config.EXCEL_DIR, "Summary_Report.xlsx")
        wb_sum.save(summary_path)
        logger.info("All 4 Excel reports generated successfully.")
